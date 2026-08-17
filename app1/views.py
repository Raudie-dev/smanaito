from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.hashers import check_password, make_password
from django.db.models import Q, Sum
from django.http import JsonResponse, HttpResponse
import datetime
import openpyxl
from .models import User, Finca, Animal, Rebaño, RegistroOrdeno, ConfiguracionUsuario, VentaAnimal, PlanVacunacion, IncidenteSanitario, GastoFinca, PrecioLecheConfig, LogActividad, Corral, PesajeAnimal, RegistroAlimentacion, TareaDiaria, HistorialTransferencia

def registrar_log(usuario, finca_id, accion, modulo, descripcion):
    try:
        LogActividad.objects.create(
            usuario=usuario,
            finca_id=finca_id,
            accion=accion,
            modulo=modulo,
            descripcion=descripcion
        )
    except Exception:
        pass

def login(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        password = request.POST.get('password', '')

        try:
            user = User.objects.get(nombre=nombre)
            if user.bloqueado:
                messages.error(request, 'Usuario bloqueado')
            elif check_password(password, user.password):
                request.session['user'] = user.id
                ConfiguracionUsuario.objects.get_or_create(user=user)
                registrar_log(user, None, 'LOGIN', 'SEGURIDAD', f"Inicio de sesión exitoso para {user.nombre}")
                return redirect('control')
            # Soporte temporal para contraseñas antiguas no encriptadas:
            elif not user.password.startswith('pbkdf2_') and user.password == password:
                # Encriptamos la contraseña al vuelo para el futuro
                user.password = make_password(password)
                user.save()
                request.session['user'] = user.id
                ConfiguracionUsuario.objects.get_or_create(user=user)
                registrar_log(user, None, 'LOGIN', 'SEGURIDAD', f"Inicio de sesión exitoso para {user.nombre} (encriptada al vuelo)")
                return redirect('control')
            else:
                messages.error(request, 'Contraseña incorrecta')
            return render(request, 'login.html')
        except User.DoesNotExist:
            messages.error(request, 'Usuario no encontrado')
            return render(request, 'login.html')

    return render(request, 'login.html')

def signup(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        email = request.POST.get('email')
        password = request.POST.get('password')
        password_confirm = request.POST.get('password_confirm')
        terminos = request.POST.get('terminos')
        
        if not terminos:
            messages.error(request, 'Debes aceptar los términos y condiciones')
            return redirect('signup')

        if password != password_confirm:
            messages.error(request, 'Las contraseñas no coinciden')
            return redirect('signup')

        if User.objects.filter(nombre=nombre).exists():
            messages.error(request, 'Este nombre de usuario ya está registrado')
            return redirect('signup')
            
        if email and User.objects.filter(email=email).exists():
            messages.error(request, 'Este correo electrónico ya está registrado')
            return redirect('signup')
            
        # Crear Usuario en app1 con contraseña encriptada
        hashed_password = make_password(password)
        nuevo_user = User.objects.create(nombre=nombre, email=email, password=hashed_password)
        registrar_log(nuevo_user, None, 'CREACION', 'SEGURIDAD', f"Nuevo usuario registrado: {nuevo_user.nombre}")
        
        # Crear Suscripcion Trial en app2 automáticamente (15 días)
        try:
            from app2.models import Suscripcion
            import datetime
            Suscripcion.objects.create(
                usuario=nuevo_user,
                plan='TRIAL',
                estado='ACTIVA',
                fecha_vencimiento=datetime.date.today() + datetime.timedelta(days=15)
            )
        except Exception as e:
            # Si falla app2 por alguna razón, no bloqueamos el signup pero el panel no lo verá bien
            print("Error creando suscripción SaaS:", e)

        messages.success(request, 'Cuenta creada exitosamente. Inicie sesión para comenzar.')
        return redirect('login')

    return render(request, 'signup.html')

def index(request):
    return render(request, 'index.html')

def cambiar_finca(request):
    if request.method == 'POST':
        finca_id = request.POST.get('finca_id')
        request.session['finca_activa_id'] = int(finca_id)
        return redirect(request.META.get('HTTP_REFERER', 'control'))
    return redirect('control')

def crear_finca(request):
    user_id = request.session.get('user')
    if not user_id:
        return redirect('login')
    user = User.objects.get(id=user_id)
    
    if request.method == 'POST':
        try:
            plan = user.suscripcion_saas.plan
        except:
            plan = 'TRIAL'
            
        fincas_count = user.fincas.count()
        if plan in ['TRIAL', 'BASICO']:
            limite = 1
        elif plan == 'PLUS':
            limite = 3
        else: # PREMIUM, VIP
            limite = float('inf')
        
        if fincas_count >= limite:
            messages.error(request, f'Límite de fincas alcanzado para su plan {plan}. ¡Contacte a soporte para un Upgrade!')
        else:
            nombre = request.POST.get('nombre')
            f = Finca.objects.create(usuario=user, nombre=nombre)
            request.session['finca_activa_id'] = f.id
            messages.success(request, f'Finca {nombre} creada exitosamente.')
            
    return redirect(request.META.get('HTTP_REFERER', 'control'))

def get_finca_context(request, user):
    fincas_usuario = user.fincas.all()
    finca_activa_id = request.session.get('finca_activa_id')
    
    if not finca_activa_id and fincas_usuario.exists():
        finca_activa_id = fincas_usuario.first().id
        request.session['finca_activa_id'] = finca_activa_id
        
    return finca_activa_id, fincas_usuario

def control(request):
    user_id = request.session.get('user')
    if not user_id:
        messages.error(request, 'Debe iniciar sesión primero')
        return redirect('login')
        
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    
    mostrar_modal_bienvenida = False
    if not finca_activa_id:
        mostrar_modal_bienvenida = True
        fincas_usuario = user.fincas.all()

    config, _ = ConfiguracionUsuario.objects.get_or_create(user=user)
    
    # 1. Total Animales
    total_animales = Animal.objects.filter(finca_id=finca_activa_id).count()
    total_machos = Animal.objects.filter(finca_id=finca_activa_id, sexo='M').count()
    total_hembras = Animal.objects.filter(finca_id=finca_activa_id, sexo='H').count()
    
    # 2. Producción Leche (Semanal)
    hace_7_dias = datetime.date.today() - datetime.timedelta(days=7)
    prod_semanal = RegistroOrdeno.objects.filter(finca_id=finca_activa_id, fecha__gte=hace_7_dias).aggregate(total=Sum('cantidad_litros'))['total'] or 0
    
    # Producción de leche diaria de los últimos 7 días para gráfico
    dias_leche = []
    valores_leche = []
    hoy = datetime.date.today()
    for i in range(6, -1, -1):
        dia = hoy - datetime.timedelta(days=i)
        dias_leche.append(dia.strftime("%d %b"))
        cant = RegistroOrdeno.objects.filter(finca_id=finca_activa_id, fecha=dia).aggregate(total=Sum('cantidad_litros'))['total'] or 0
        valores_leche.append(float(cant))
        
    # 3. Alertas Destete
    umbral_destete = datetime.date.today() - datetime.timedelta(days=config.meses_destete * 30)
    alertas_destete_count = Animal.objects.filter(finca_id=finca_activa_id, destetado=False, fecha_nacimiento__lte=umbral_destete).count() if config.usar_destete else 0
    
    # 4. Indicadores Financieros Rápidos
    precio_leche_config, _ = PrecioLecheConfig.objects.get_or_create(finca_id=finca_activa_id, defaults={'precio_por_litro': 0.00})
    litros_totales = RegistroOrdeno.objects.filter(finca_id=finca_activa_id).aggregate(total=Sum('cantidad_litros'))['total'] or 0
    ingresos_leche = float(litros_totales) * float(precio_leche_config.precio_por_litro)
    ingresos_animales = VentaAnimal.objects.filter(finca_id=finca_activa_id).aggregate(total=Sum('precio_total'))['total'] or 0
    egresos_totales = GastoFinca.objects.filter(finca_id=finca_activa_id).aggregate(total=Sum('monto'))['total'] or 0
    balance_neto = (ingresos_leche + float(ingresos_animales)) - float(egresos_totales)

    # 5. Últimos Eventos
    eventos = []
    
    ultimos_ordenos = RegistroOrdeno.objects.filter(finca_id=finca_activa_id).order_by('-fecha', '-hora_finalizacion')[:3]
    for o in ultimos_ordenos:
        eventos.append({
            'id': f"#{o.rebaño.nombre[:3].upper()}-{o.id}",
            'tipo': 'Registro de Ordeño',
            'fecha': o.fecha.strftime("%d %b %Y"),
            'estado': f"{o.cantidad_litros} L",
            'badge': 'bg-success'
        })
        
    ultimos_animales = Animal.objects.filter(finca_id=finca_activa_id).order_by('-id')[:3]
    for a in ultimos_animales:
        eventos.append({
            'id': f"{a.codigo}",
            'tipo': 'Nuevo Ingreso',
            'fecha': 'Reciente',
            'estado': 'Activo',
            'badge': 'bg-primary'
        })
        
    context = {
        'total_animales': total_animales,
        'total_machos': total_machos,
        'total_hembras': total_hembras,
        'prod_semanal': prod_semanal,
        'alertas_destete_count': alertas_destete_count,
        'balance_neto': balance_neto,
        'dias_leche': dias_leche,
        'valores_leche': valores_leche,
        'eventos': eventos[:6],
        'config': config,
        'fincas_usuario': fincas_usuario,
        'mostrar_modal_bienvenida': mostrar_modal_bienvenida,
    }
        
    return render(request, 'control.html', context)

def registro(request):
    user_id = request.session.get('user')
    if not user_id:
        messages.error(request, 'Debe iniciar sesión primero')
        return redirect('login')
        
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    
    if not finca_activa_id:
        messages.error(request, 'Debe crear una finca primero')
        return redirect('control')

    if request.method == 'POST':
        animal_id = request.POST.get('animal_id')
        codigo = request.POST.get('codigo')
        nombre = request.POST.get('nombre')
        propietario = request.POST.get('propietario', '')
        fecha_nacimiento = request.POST.get('fecha_nacimiento')
        sexo = request.POST.get('sexo')
        estado_gestacion = request.POST.get('estado_gestacion', 'N_A')
        estado_produccion = request.POST.get('estado_produccion', 'N_A')
        uso_macho = request.POST.get('uso_macho', 'N_A')
        rebaño_id = request.POST.get('rebaño')
        papa_codigo = request.POST.get('papa')
        mama_codigo = request.POST.get('mama')
        
        try:
            rebaño_obj = Rebaño.objects.get(id=rebaño_id, finca_id=finca_activa_id) if rebaño_id else None
            
            # Gestión de Padres (Si no existen, se crean al vuelo)
            papa = None
            if papa_codigo:
                papa, _ = Animal.objects.get_or_create(
                    finca_id=finca_activa_id,
                    codigo=papa_codigo,
                    defaults={
                        'nombre': f"Padre Externo {papa_codigo}",
                        'sexo': 'M',
                        'uso_macho': 'REPRODUCTOR',
                        'fecha_nacimiento': datetime.date.today() - datetime.timedelta(days=365*4)
                    }
                )
                
            mama = None
            if mama_codigo:
                mama, _ = Animal.objects.get_or_create(
                    finca_id=finca_activa_id,
                    codigo=mama_codigo,
                    defaults={
                        'nombre': f"Madre Externa {mama_codigo}",
                        'sexo': 'H',
                        'fecha_nacimiento': datetime.date.today() - datetime.timedelta(days=365*4)
                    }
                )
            
            if animal_id:
                # Edición
                anim = Animal.objects.get(id=animal_id, finca_id=finca_activa_id)
                anim.codigo = codigo
                anim.nombre = nombre
                anim.propietario = propietario
                anim.fecha_nacimiento = fecha_nacimiento
                anim.sexo = sexo
                anim.estado_gestacion = estado_gestacion
                anim.estado_produccion = estado_produccion
                anim.uso_macho = uso_macho
                anim.rebaño = rebaño_obj
                anim.papa = papa
                anim.mama = mama
                anim.save()
                registrar_log(user, finca_activa_id, 'MODIFICACION', 'ANIMALES', f"Actualizó datos del animal: '{nombre}' ({codigo})")
                messages.success(request, f"Animal {nombre} actualizado exitosamente.")
            else:
                # Creación
                Animal.objects.create(
                    finca_id=finca_activa_id,
                    codigo=codigo,
                    nombre=nombre,
                    propietario=propietario,
                    fecha_nacimiento=fecha_nacimiento,
                    sexo=sexo,
                    estado_gestacion=estado_gestacion,
                    estado_produccion=estado_produccion,
                    uso_macho=uso_macho,
                    rebaño=rebaño_obj,
                    papa=papa,
                    mama=mama
                )
                registrar_log(user, finca_activa_id, 'CREACION', 'ANIMALES', f"Registró nuevo animal: '{nombre}' ({codigo})")
                messages.success(request, f"Animal {nombre} registrado exitosamente.")
                
            return redirect('registro')
        except Exception as e:
            messages.error(request, f"Error al registrar animal: {str(e)}")
            
    # Lógica GET y Filtros
    q = request.GET.get('q', '')
    filtro_sexo = request.GET.get('sexo', '')
    filtro_rebano = request.GET.get('rebano', '')
    
    if filtro_rebano:
        try:
            r = Rebaño.objects.get(id=filtro_rebano, finca_id=finca_activa_id)
            animales = r.get_animales()
        except Rebaño.DoesNotExist:
            animales = Animal.objects.filter(finca_id=finca_activa_id)
    else:
        animales = Animal.objects.filter(finca_id=finca_activa_id)
    
    if q:
        animales = animales.filter(Q(nombre__icontains=q) | Q(codigo__icontains=q))
    if filtro_sexo:
        animales = animales.filter(sexo=filtro_sexo)
        
    propietarios_unicos = Animal.objects.filter(finca_id=finca_activa_id).exclude(propietario__isnull=True).exclude(propietario='').values_list('propietario', flat=True).distinct()
        
    context = {
        'animales': animales,
        'rebanos': Rebaño.objects.filter(finca_id=finca_activa_id),
        'q': q,
        'filtro_sexo': filtro_sexo,
        'filtro_rebano': filtro_rebano,
        'fincas_usuario': fincas_usuario,
        'propietarios_unicos': propietarios_unicos,
    }
    return render(request, 'registro.html', context)


def rebaño(request):
    user_id = request.session.get('user')
    if not user_id:
        messages.error(request, 'Debe iniciar sesión primero')
        return redirect('login')
        
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    if not finca_activa_id:
        messages.error(request, 'Debe crear una finca primero')
        return redirect('control')
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'generar_estandares':
            Rebaño.objects.get_or_create(finca_id=finca_activa_id, nombre='Rebaño Ordeño', defaults={
                'descripcion': 'Vacas en Lactancia',
                'es_dinamico': True, 'filtro_sexo': 'H', 'filtro_estado_produccion': 'LACTANCIA'
            })
            Rebaño.objects.get_or_create(finca_id=finca_activa_id, nombre='Rebaño Horro', defaults={
                'descripcion': 'Vacas Secas (Preñadas o Vacías)',
                'es_dinamico': True, 'filtro_sexo': 'H', 'filtro_estado_produccion': 'SECA'
            })
            Rebaño.objects.get_or_create(finca_id=finca_activa_id, nombre='Rebaño Engorde', defaults={
                'descripcion': 'Machos para sacrificio',
                'es_dinamico': True, 'filtro_sexo': 'M', 'filtro_uso_macho': 'ENGORDE'
            })
            Rebaño.objects.get_or_create(finca_id=finca_activa_id, nombre='Rebaño Mautes', defaults={
                'descripcion': 'Animales de más de 1 año',
                'es_dinamico': True, 'filtro_edad_min_meses': 12
            })
            messages.success(request, "Rebaños estándar generados exitosamente.")
        else:
            # Crear o Editar
            rebano_id = request.POST.get('rebano_id')
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion')
            es_dinamico = request.POST.get('es_dinamico') == 'on'
            filtro_sexo = request.POST.get('filtro_sexo') or None
            
            # Edades
            min_meses = request.POST.get('filtro_edad_min_meses')
            max_meses = request.POST.get('filtro_edad_max_meses')
            min_meses = int(min_meses) if min_meses else None
            max_meses = int(max_meses) if max_meses else None
            
            # Estados
            f_gestacion = request.POST.get('filtro_estado_gestacion') or None
            f_produccion = request.POST.get('filtro_estado_produccion') or None
            f_uso = request.POST.get('filtro_uso_macho') or None
            
            try:
                if rebano_id:
                    # Editar
                    reb = Rebaño.objects.get(id=rebano_id, finca_id=finca_activa_id)
                    reb.nombre = nombre
                    reb.descripcion = descripcion
                    reb.es_dinamico = es_dinamico
                    reb.filtro_sexo = filtro_sexo
                    reb.filtro_edad_min_meses = min_meses
                    reb.filtro_edad_max_meses = max_meses
                    reb.filtro_estado_gestacion = f_gestacion
                    reb.filtro_estado_produccion = f_produccion
                    reb.filtro_uso_macho = f_uso
                    reb.save()
                    messages.success(request, f"Rebaño '{nombre}' actualizado exitosamente.")
                else:
                    # Crear
                    Rebaño.objects.create(
                        finca_id=finca_activa_id,
                        nombre=nombre, 
                        descripcion=descripcion,
                        es_dinamico=es_dinamico,
                        filtro_sexo=filtro_sexo,
                        filtro_edad_min_meses=min_meses,
                        filtro_edad_max_meses=max_meses,
                        filtro_estado_gestacion=f_gestacion,
                        filtro_estado_produccion=f_produccion,
                        filtro_uso_macho=f_uso
                    )
                    messages.success(request, f"Rebaño '{nombre}' creado exitosamente.")
            except Exception as e:
                messages.error(request, f"Error al guardar rebaño: {str(e)}")
                
        return redirect('rebaño')
            
    rebanos = Rebaño.objects.filter(finca_id=finca_activa_id)
    for r in rebanos:
        r.cantidad_animales = r.count_animales()
        
    context = {
        'rebanos': rebanos,
        'fincas_usuario': fincas_usuario,
    }
    return render(request, 'rebaño.html', context)

def ordeño(request):
    user_id = request.session.get('user')
    if not user_id:
        messages.error(request, 'Debe iniciar sesión primero')
        return redirect('login')
        
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    if not finca_activa_id:
        messages.error(request, 'Debe crear una finca primero')
        return redirect('control')
    # Obtener el rebaño activo (por session o GET)
    rebano_activo_id = request.GET.get('rebano')
    if rebano_activo_id:
        request.session['rebano_ordeno_activo'] = rebano_activo_id
    else:
        rebano_activo_id = request.session.get('rebano_ordeno_activo')
        
    rebano_activo = None
    if rebano_activo_id:
        try:
            rebano_activo = Rebaño.objects.get(id=rebano_activo_id, finca_id=finca_activa_id)
        except Rebaño.DoesNotExist:
            rebano_activo = None
            
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'registrar_ordeno':
            rebano_post = request.POST.get('rebano_id')
            litros = request.POST.get('cantidad_litros')
            hora = request.POST.get('hora_finalizacion')
            obs = request.POST.get('observaciones', '')
            
            try:
                r = Rebaño.objects.get(id=rebano_post, finca_id=finca_activa_id)
                
                if r.count_animales() == 0:
                    messages.error(request, f"El rebaño '{r.nombre}' no tiene animales registrados. No se puede registrar producción de ordeño.")
                else:
                    RegistroOrdeno.objects.create(
                        finca_id=finca_activa_id,
                        rebaño=r,
                        cantidad_litros=litros,
                        hora_finalizacion=hora,
                        observaciones=obs
                    )
                    registrar_log(user, finca_activa_id, 'CREACION', 'ORDEÑO', f"Registró producción de ordeño: {litros} L para el rebaño '{r.nombre}'")
                    messages.success(request, f"Producción de {litros}L registrada exitosamente para el rebaño {r.nombre}.")

            except Exception as e:
                messages.error(request, f"Error al registrar ordeño: {str(e)}")
                
        elif action == 'agregar_vaca':
            codigo_vaca = request.POST.get('codigo_vaca')
            if codigo_vaca:
                try:
                    vaca = Animal.objects.get(codigo=codigo_vaca, sexo='H', finca_id=finca_activa_id)
                    vaca.estado_produccion = 'LACTANCIA'
                    
                    # Si el rebaño activo NO es dinámico, podemos asociar la vaca a este rebaño fijo.
                    if rebano_activo and not rebano_activo.es_dinamico:
                        vaca.rebaño = rebano_activo
                        
                    vaca.save()
                    registrar_log(user, finca_activa_id, 'MODIFICACION', 'ANIMALES', f"Activó estado de Lactancia para la vaca '{vaca.nombre}' ({vaca.codigo})")
                    messages.success(request, f"Vaca {vaca.nombre} ({vaca.codigo}) ha sido pasada a estado 'En Lactancia'.")
                except Animal.DoesNotExist:
                    messages.error(request, "Vaca no encontrada o no es hembra.")
                    
        return redirect('ordeño')
        
    # Obtener datos para la vista (solo rebaños con característica de ordeño/lactancia)
    rebanos = Rebaño.objects.filter(finca_id=finca_activa_id, filtro_estado_produccion='LACTANCIA')
    
    # Historial de ordeños de los últimos 30 días
    hace_30_dias = datetime.date.today() - datetime.timedelta(days=30)
    
    qs_registros = RegistroOrdeno.objects.filter(finca_id=finca_activa_id, fecha__gte=hace_30_dias).order_by('-fecha', '-hora_finalizacion')
    if rebano_activo:
        qs_registros = qs_registros.filter(rebaño=rebano_activo)
        
    total_30_dias = qs_registros.aggregate(total=Sum('cantidad_litros'))['total'] or 0
        
    context = {
        'rebanos': rebanos,
        'rebano_activo': rebano_activo,
        'registros': qs_registros[:20], # Mostrar últimos 20
        'total_30_dias': total_30_dias,
        'fincas_usuario': fincas_usuario,
    }
    return render(request, 'ordeño.html', context)

def crianza(request):
    user_id = request.session.get('user')
    if not user_id:
        messages.error(request, 'Debe iniciar sesión primero')
        return redirect('login')
        
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    if not finca_activa_id:
        messages.error(request, 'Debe crear una finca primero')
        return redirect('control')
    config, _ = ConfiguracionUsuario.objects.get_or_create(user=user)
        
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'guardar_configuracion':
            meses_destete = request.POST.get('meses_destete')
            meses_mamanto = request.POST.get('meses_mamanto')
            usar_mamanto = request.POST.get('usar_mamanto') == 'on'
            usar_destete = request.POST.get('usar_destete') == 'on'
            
            config.meses_destete = int(meses_destete) if meses_destete else 7
            config.meses_mamanto = int(meses_mamanto) if meses_mamanto else 3
            config.usar_mamanto = usar_mamanto
            config.usar_destete = usar_destete
            config.save()
            
            messages.success(request, "Configuración de reglas de crianza actualizada.")
            
        elif action == 'destetar_animal':
            animal_id = request.POST.get('animal_id')
            nuevo_rebano_id = request.POST.get('rebano_id')
            
            try:
                animal = Animal.objects.get(id=animal_id, finca_id=finca_activa_id)
                animal.destetado = True
                
                if nuevo_rebano_id:
                    reb = Rebaño.objects.get(id=nuevo_rebano_id, finca_id=finca_activa_id)
                    animal.rebaño = reb
                    
                animal.save()
                messages.success(request, f"El animal {animal.codigo} ha sido destetado exitosamente.")
            except Exception as e:
                messages.error(request, f"Error al destetar: {str(e)}")
                
        return redirect('crianza')

    # Calcular fechas umbrales
    umbral_destete = datetime.date.today() - datetime.timedelta(days=config.meses_destete * 30)
    umbral_mamanto = datetime.date.today() - datetime.timedelta(days=config.meses_mamanto * 30)
    
    no_destetados = Animal.objects.filter(finca_id=finca_activa_id, destetado=False)
    mamanto = []
    alertas_destete = []
    today = datetime.date.today()
    
    if config.usar_mamanto:
        mamanto = no_destetados.filter(fecha_nacimiento__gt=umbral_mamanto).order_by('-fecha_nacimiento')
        for b in mamanto:
            b.edad_meses = (today - b.fecha_nacimiento).days // 30
            
    if config.usar_destete:
        alertas_destete = no_destetados.filter(fecha_nacimiento__lte=umbral_destete).order_by('fecha_nacimiento')
        for b in alertas_destete:
            b.edad_meses = (today - b.fecha_nacimiento).days // 30

    rebanos = Rebaño.objects.filter(finca_id=finca_activa_id)

    context = {
        'config': config,
        'mamanto': mamanto,
        'alertas_destete': alertas_destete,
        'rebanos': rebanos,
        'fincas_usuario': fincas_usuario,
    }
    
    return render(request, 'crianza.html', context)

def ventas(request):
    user_id = request.session.get('user')
    if not user_id:
        messages.error(request, 'Debe iniciar sesión primero')
        return redirect('login')
        
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    if not finca_activa_id:
        messages.error(request, 'Debe crear una finca primero')
        return redirect('control')
        
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'registrar_venta':
            animal_ids = request.POST.getlist('animal_ids')
            motivo = request.POST.get('motivo')
            kilos = request.POST.get('kilos')
            precio = request.POST.get('precio_total')
            comprador = request.POST.get('comprador')
            obs = request.POST.get('observaciones')
            
            if not animal_ids:
                messages.error(request, "Debe seleccionar al menos un animal.")
                return redirect('ventas')
                
            try:
                venta = VentaAnimal.objects.create(
                    usuario=user,
                    finca_id=finca_activa_id,
                    motivo=motivo,
                    kilos=kilos if kilos else None,
                    precio_total=precio if precio else None,
                    comprador=comprador,
                    observaciones=obs
                )
                
                for a_id in animal_ids:
                    animal = Animal.objects.get(id=a_id, finca_id=finca_activa_id, estado_vida='VIVO')
                    animal.estado_vida = 'VENDIDO'
                    animal.rebaño = None
                    animal.save()
                    venta.animales.add(animal)
                    
                registrar_log(user, finca_activa_id, 'CREACION', 'VENTAS', f"Registró la venta de {len(animal_ids)} animal(es) por un monto total de ${precio}")
                messages.success(request, f"La venta de {len(animal_ids)} animal(es) ha sido registrada con éxito.")
            except Exception as e:
                messages.error(request, f"Error al registrar la venta: {str(e)}")
                
        return redirect('ventas')
        
    animales_vivos = Animal.objects.filter(finca_id=finca_activa_id, estado_vida='VIVO')
    ventas_historico = VentaAnimal.objects.filter(finca_id=finca_activa_id).order_by('-fecha_venta')
    
    total_vendidos = ventas_historico.count()
    total_kilos = sum(v.kilos for v in ventas_historico if v.kilos)
    
    context = {
        'animales_vivos': animales_vivos,
        'ventas': ventas_historico,
        'fincas_usuario': fincas_usuario,
        'total_vendidos': total_vendidos,
        'total_kilos': total_kilos,
    }
    return render(request, 'ventas.html', context)

def descargar_plantilla(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Animales"
    
    headers = ['Codigo', 'Nombre', 'Sexo', 'Fecha_Nacimiento', 'Propietario']
    ws.append(headers)
    
    ejemplo = ['VAC-001', 'Lola', 'H', '2023-05-14', 'Mi Finca']
    ws.append(ejemplo)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_animales.xlsx"'
    wb.save(response)
    return response

def datos_animales(request):
    user_id = request.session.get('user')
    if not user_id:
        messages.error(request, 'Debe iniciar sesión primero')
        return redirect('login')
        
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    if not finca_activa_id:
        messages.error(request, 'Debe crear una finca primero')
        return redirect('control')
        
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'importar_excel':
            archivo = request.FILES.get('archivo_excel')
            if not archivo:
                messages.error(request, 'Por favor seleccione un archivo Excel.')
            elif not archivo.name.endswith('.xlsx'):
                messages.error(request, 'El formato del archivo debe ser .xlsx')
            else:
                try:
                    wb = openpyxl.load_workbook(archivo)
                    ws = wb.active
                    
                    headers = [str(cell.value).strip().lower() for cell in ws[1]]
                    
                    if 'codigo' not in headers or 'nombre' not in headers or 'sexo' not in headers or 'fecha_nacimiento' not in headers:
                        messages.error(request, 'El Excel no tiene el formato correcto. Use la plantilla.')
                    else:
                        idx_codigo = headers.index('codigo')
                        idx_nombre = headers.index('nombre')
                        idx_sexo = headers.index('sexo')
                        idx_fecha = headers.index('fecha_nacimiento')
                        idx_prop = headers.index('propietario') if 'propietario' in headers else -1
                        
                        creados = 0
                        omitidos = 0
                        
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if not row[idx_codigo] or not row[idx_nombre]:
                                continue
                                
                            codigo = str(row[idx_codigo]).strip()
                            nombre = str(row[idx_nombre]).strip()
                            sexo_raw = str(row[idx_sexo]).strip().upper()
                            sexo = 'H' if 'H' in sexo_raw or 'F' in sexo_raw else 'M'
                            
                            fecha = row[idx_fecha]
                            if isinstance(fecha, datetime.datetime):
                                fecha_nac = fecha.date()
                            else:
                                try:
                                    fecha_str = str(fecha).strip()
                                    fecha_nac = datetime.datetime.strptime(fecha_str, "%Y-%m-%d").date()
                                except:
                                    fecha_nac = datetime.date.today()
                                    
                            prop = str(row[idx_prop]).strip() if idx_prop >= 0 and row[idx_prop] else None
                            
                            if not Animal.objects.filter(finca_id=finca_activa_id, codigo=codigo).exists():
                                Animal.objects.create(
                                    usuario=user,
                                    finca_id=finca_activa_id,
                                    codigo=codigo,
                                    nombre=nombre,
                                    sexo=sexo,
                                    fecha_nacimiento=fecha_nac,
                                    propietario=prop
                                )
                                creados += 1
                            else:
                                omitidos += 1
                                
                        registrar_log(user, finca_activa_id, 'IMPORTACION', 'ANIMALES', f"Importó animales desde Excel: {creados} creados, {omitidos} omitidos")
                        messages.success(request, f"Importación exitosa. Creados: {creados}. Omitidos (código existente): {omitidos}.")
                except Exception as e:
                    messages.error(request, f"Error al procesar el Excel: {str(e)}")
            
            return redirect('datos_animales')

    total_animales = Animal.objects.filter(finca_id=finca_activa_id, estado_vida='VIVO').count()
    context = {
        'fincas_usuario': fincas_usuario,
        'total_animales': total_animales,
    }
    return render(request, 'datos_animales.html', context)


def vacunacion(request):
    user_id = request.session.get('user')
    if not user_id:
        messages.error(request, 'Debe iniciar sesión primero')
        return redirect('login')
        
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    if not finca_activa_id:
        messages.error(request, 'Debe crear una finca primero')
        return redirect('control')
        
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'programar_vacuna':
            vacuna = request.POST.get('vacuna')
            fecha_prog = request.POST.get('fecha_programada')
            rebano_id = request.POST.get('rebano_id')
            observaciones = request.POST.get('observaciones')
            
            try:
                reb = Rebaño.objects.get(id=rebano_id, finca_id=finca_activa_id) if rebano_id else None
                PlanVacunacion.objects.create(
                    usuario=user,
                    finca_id=finca_activa_id,
                    vacuna=vacuna,
                    fecha_programada=fecha_prog,
                    rebaño=reb,
                    observaciones=observaciones
                )
                registrar_log(user, finca_activa_id, 'CREACION', 'SANIDAD', f"Programó vacuna: '{vacuna}' para el {fecha_prog}")
                messages.success(request, f"Plan de vacunación para {vacuna} programado con éxito.")
            except Exception as e:
                messages.error(request, f"Error al programar: {str(e)}")
                
        elif action == 'completar_vacuna':
            vacuna_id = request.POST.get('vacuna_id')
            try:
                plan = PlanVacunacion.objects.get(id=vacuna_id, finca_id=finca_activa_id)
                plan.estado = 'COMPLETADO'
                plan.fecha_aplicacion = datetime.date.today()
                plan.save()
                registrar_log(user, finca_activa_id, 'MODIFICACION', 'SANIDAD', f"Marcó como completada la vacunación: '{plan.vacuna}'")
                messages.success(request, f"Vacunación {plan.vacuna} marcada como completada.")
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
                
        return redirect('vacunacion')
        
    pendientes = PlanVacunacion.objects.filter(finca_id=finca_activa_id, estado='PENDIENTE').order_by('fecha_programada')
    completadas = PlanVacunacion.objects.filter(finca_id=finca_activa_id, estado='COMPLETADO').order_by('-fecha_aplicacion')
    rebaños = Rebaño.objects.filter(finca_id=finca_activa_id)
    
    context = {
        'fincas_usuario': fincas_usuario,
        'pendientes': pendientes,
        'completadas': completadas,
        'rebaños': rebaños,
    }
    return render(request, 'vacunacion.html', context)


def incidentes(request):
    user_id = request.session.get('user')
    if not user_id:
        messages.error(request, 'Debe iniciar sesión primero')
        return redirect('login')
        
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    if not finca_activa_id:
        messages.error(request, 'Debe crear una finca primero')
        return redirect('control')
        
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'registrar_incidente':
            animal_id = request.POST.get('animal_id')
            tipo = request.POST.get('tipo')
            fecha = request.POST.get('fecha_incidente')
            diag = request.POST.get('diagnostico')
            trat = request.POST.get('tratamiento')
            
            try:
                animal = Animal.objects.get(id=animal_id, finca_id=finca_activa_id)
                IncidenteSanitario.objects.create(
                    usuario=user,
                    finca_id=finca_activa_id,
                    animal=animal,
                    fecha_incidente=fecha,
                    tipo=tipo,
                    diagnostico=diag,
                    tratamiento=trat
                )
                registrar_log(user, finca_activa_id, 'CREACION', 'SANIDAD', f"Registró incidente clínico para el animal '{animal.codigo}': {diag}")
                messages.success(request, f"Incidente registrado para el animal {animal.codigo}.")
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
                
        elif action == 'resolver_incidente':
            incidente_id = request.POST.get('incidente_id')
            try:
                inc = IncidenteSanitario.objects.get(id=incidente_id, finca_id=finca_activa_id)
                inc.estado = 'RESUELTO'
                inc.save()
                registrar_log(user, finca_activa_id, 'MODIFICACION', 'SANIDAD', f"Marcó como resuelto el incidente clínico del animal '{inc.animal.codigo}'")
                messages.success(request, f"Incidente de {inc.animal.codigo} resuelto.")
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
                
        return redirect('incidentes')
        
    activos = IncidenteSanitario.objects.filter(finca_id=finca_activa_id, estado='ACTIVO').order_by('-fecha_incidente')
    resueltos = IncidenteSanitario.objects.filter(finca_id=finca_activa_id, estado='RESUELTO').order_by('-fecha_incidente')[:50]
    animales_vivos = Animal.objects.filter(finca_id=finca_activa_id, estado_vida='VIVO')
    
    context = {
        'fincas_usuario': fincas_usuario,
        'activos': activos,
        'resueltos': resueltos,
        'animales_vivos': animales_vivos,
    }
    return render(request, 'incidentes.html', context)


# Endpoints API
def api_buscar_padre(request):
    user_id = request.session.get('user')
    if not user_id:
        return JsonResponse({'results': []})
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    if not finca_activa_id:
        messages.error(request, 'Debe crear una finca primero')
        return redirect('control')
    
    q = request.GET.get('q', '')
    if len(q) < 1:
        return JsonResponse({'results': []})
        
    # Machos reproductores
    machos = Animal.objects.filter(finca_id=finca_activa_id, sexo='M', uso_macho='REPRODUCTOR').filter(
        Q(codigo__icontains=q) | Q(nombre__icontains=q)
    )[:10]
    
    results = [{'codigo': m.codigo, 'nombre': m.nombre} for m in machos]
    return JsonResponse({'results': results})

def api_buscar_madre(request):
    user_id = request.session.get('user')
    if not user_id:
        return JsonResponse({'results': []})
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    if not finca_activa_id:
        messages.error(request, 'Debe crear una finca primero')
        return redirect('control')

    q = request.GET.get('q', '')
    if len(q) < 1:
        return JsonResponse({'results': []})
        
    # Hembras de >= 3 años (36 meses) aprox 1095 dias
    tres_anios = datetime.date.today() - datetime.timedelta(days=1095)
    hembras = Animal.objects.filter(finca_id=finca_activa_id, sexo='H', fecha_nacimiento__lte=tres_anios).filter(
        Q(codigo__icontains=q) | Q(nombre__icontains=q)
    )[:10]
    
    results = [{'codigo': h.codigo, 'nombre': h.nombre} for h in hembras]
    return JsonResponse({'results': results})

def api_buscar_vaca_seca(request):
    user_id = request.session.get('user')
    if not user_id:
        return JsonResponse({'results': []})
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    if not finca_activa_id:
        messages.error(request, 'Debe crear una finca primero')
        return redirect('control')

    q = request.GET.get('q', '')
    if len(q) < 1:
        return JsonResponse({'results': []})
        
    # Buscar hembras que NO están en lactancia
    vacas = Animal.objects.filter(finca_id=finca_activa_id, sexo='H', estado_produccion='SECA').filter(
        Q(codigo__icontains=q) | Q(nombre__icontains=q)
    )[:10]
    
    results = [
        {'id': a.id, 'text': f"{a.codigo} - {a.nombre}"}
        for a in vacas
    ]
    return JsonResponse({'results': results})

def api_buscar_animal_vivo(request):
    user_id = request.session.get('user')
    if not user_id:
        return JsonResponse({'results': []})
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    if not finca_activa_id:
        return JsonResponse({'results': []})
        
    q = request.GET.get('q', '').strip()
    if len(q) < 1:
        return JsonResponse({'results': []})
        
    animales = Animal.objects.filter(
        finca_id=finca_activa_id, 
        estado_vida='VIVO'
    ).filter(
        Q(codigo__icontains=q) | Q(nombre__icontains=q)
    )[:15]
    
    results = [
        {'id': a.id, 'text': f"{a.codigo} - {a.nombre}"}
        for a in animales
    ]
    return JsonResponse({'results': results})

def api_reportes(request):
    user_id = request.session.get('user')
    if not user_id:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
        
    user = User.objects.get(id=user_id)
    finca_activa_id, _ = get_finca_context(request, user)
    
    if not finca_activa_id:
        return JsonResponse({'error': 'No finca'}, status=400)
        
    sexo = request.GET.get('sexo')
    estado_prod = request.GET.get('estado_produccion')
    estado_gest = request.GET.get('estado_gestacion')
    uso_macho = request.GET.get('uso_macho')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    qs = Animal.objects.filter(finca_id=finca_activa_id, estado_vida='VIVO')
    
    if sexo:
        qs = qs.filter(sexo=sexo)
    if estado_prod:
        qs = qs.filter(estado_produccion=estado_prod)
    if estado_gest:
        qs = qs.filter(estado_gestacion=estado_gest)
    if uso_macho:
        qs = qs.filter(uso_macho=uso_macho)
        
    if fecha_desde:
        qs = qs.filter(fecha_nacimiento__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha_nacimiento__lte=fecha_hasta)
        
    total = qs.count()
    machos = qs.filter(sexo='M').count()
    hembras = qs.filter(sexo='H').count()
    
    lactancia = qs.filter(estado_produccion='LACTANCIA').count()
    seca = qs.filter(estado_produccion='SECA').count()
    vacia = qs.filter(estado_gestacion='VACIA').count()
    prenada = qs.filter(estado_gestacion='PREÑADA').count()
    
    # Gráficos
    grafico_sexo = [machos, hembras]
    grafico_prod = [lactancia, seca, qs.count() - lactancia - seca]
    
    # Listado limitado para tabla preview
    animales = list(qs.order_by('-id')[:50].values('codigo', 'nombre', 'sexo', 'estado_produccion'))
    
    return JsonResponse({
        'total': total,
        'grafico_sexo': grafico_sexo,
        'grafico_prod': grafico_prod,
        'animales': animales,
        'estadisticas': {
            'lactancia': lactancia,
            'seca': seca,
            'vacia': vacia,
            'prenada': prenada
        }
    })

def exportar_reporte_excel(request):
    user_id = request.session.get('user')
    if not user_id:
        return redirect('login')
        
    user = User.objects.get(id=user_id)
    finca_activa_id, _ = get_finca_context(request, user)
    if not finca_activa_id:
        return redirect('control')
        
    sexo = request.GET.get('sexo')
    estado_prod = request.GET.get('estado_produccion')
    estado_gest = request.GET.get('estado_gestacion')
    uso_macho = request.GET.get('uso_macho')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    qs = Animal.objects.filter(finca_id=finca_activa_id, estado_vida='VIVO')
    
    if sexo:
        qs = qs.filter(sexo=sexo)
    if estado_prod:
        qs = qs.filter(estado_produccion=estado_prod)
    if estado_gest:
        qs = qs.filter(estado_gestacion=estado_gest)
    if uso_macho:
        qs = qs.filter(uso_macho=uso_macho)
        
    if fecha_desde:
        qs = qs.filter(fecha_nacimiento__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha_nacimiento__lte=fecha_hasta)
            
    # Crear archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Filtrado"
    
    # Encabezados
    headers = ['Codigo', 'Nombre', 'Sexo', 'Fecha Nacimiento', 'Estado Produccion', 'Estado Gestacion', 'Uso (Machos)']
    ws.append(headers)
    
    for a in qs.order_by('codigo'):
        ws.append([
            a.codigo, 
            a.nombre, 
            a.sexo, 
            str(a.fecha_nacimiento) if a.fecha_nacimiento else '', 
            a.estado_produccion, 
            a.estado_gestacion,
            a.uso_macho
        ])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_animales.xlsx"'
    wb.save(response)
    return response

def finanzas(request):
    user_id = request.session.get('user')
    if not user_id:
        messages.error(request, 'Debe iniciar sesión primero')
        return redirect('login')
        
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    if not finca_activa_id:
        messages.error(request, 'Debe crear una finca primero')
        return redirect('control')
        
    precio_leche_config, _ = PrecioLecheConfig.objects.get_or_create(
        finca_id=finca_activa_id,
        defaults={'precio_por_litro': 0.00}
    )
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'registrar_gasto':
            concepto = request.POST.get('concepto')
            monto = request.POST.get('monto')
            categoria = request.POST.get('categoria')
            tipo = request.POST.get('tipo', 'VARIABLE')
            fecha = request.POST.get('fecha')
            
            try:
                GastoFinca.objects.create(
                    usuario=user,
                    finca_id=finca_activa_id,
                    concepto=concepto,
                    monto=monto,
                    categoria=categoria,
                    tipo=tipo,
                    fecha=fecha
                )
                registrar_log(user, finca_activa_id, 'CREACION', 'FINANZAS', f"Registró gasto: '{concepto}' por ${monto} ({tipo})")
                messages.success(request, 'Gasto registrado correctamente')
            except Exception as e:
                messages.error(request, f'Error al registrar gasto: {str(e)}')
                
        elif action == 'configurar_leche':
            precio = request.POST.get('precio_por_litro')
            try:
                precio_leche_config.precio_por_litro = precio
                precio_leche_config.save()
                registrar_log(user, finca_activa_id, 'CONFIGURACION', 'FINANZAS', f"Actualizó el precio de la leche a ${precio} por litro")
                messages.success(request, 'Precio del litro de leche actualizado')
            except Exception as e:
                messages.error(request, f'Error al actualizar precio: {str(e)}')
                
        elif action == 'eliminar_gasto':
            gasto_id = request.POST.get('gasto_id')
            try:
                gasto = GastoFinca.objects.get(id=gasto_id, finca_id=finca_activa_id)
                concepto_del = gasto.concepto
                monto_del = gasto.monto
                gasto.delete()
                registrar_log(user, finca_activa_id, 'ELIMINACION', 'FINANZAS', f"Eliminó gasto: '{concepto_del}' por ${monto_del}")
                messages.success(request, 'Gasto eliminado')
            except Exception as e:
                messages.error(request, f'Error al eliminar: {str(e)}')
                
        return redirect('finanzas')
        
    # Cálculos Financieros
    litros_totales = RegistroOrdeno.objects.filter(finca_id=finca_activa_id).aggregate(total=Sum('cantidad_litros'))['total'] or 0
    ingresos_leche = float(litros_totales) * float(precio_leche_config.precio_por_litro)
    
    ingresos_animales = VentaAnimal.objects.filter(finca_id=finca_activa_id).aggregate(total=Sum('precio_total'))['total'] or 0
    
    egresos_totales = GastoFinca.objects.filter(finca_id=finca_activa_id).aggregate(total=Sum('monto'))['total'] or 0
    
    ingresos_totales = float(ingresos_leche) + float(ingresos_animales)
    balance_neto = ingresos_totales - float(egresos_totales)
    
    gastos = GastoFinca.objects.filter(finca_id=finca_activa_id).order_by('-fecha')
    
    # Datos para gráficos de gastos por categoría
    categorias = ['ALIMENTO', 'VETERINARIA', 'PERSONAL', 'SERVICIOS', 'MANTENIMIENTO', 'OTRO']
    chart_gastos = []
    for c in categorias:
        total_cat = GastoFinca.objects.filter(finca_id=finca_activa_id, categoria=c).aggregate(total=Sum('monto'))['total'] or 0
        chart_gastos.append(float(total_cat))
        
    context = {
        'fincas_usuario': fincas_usuario,
        'precio_leche_config': precio_leche_config,
        'litros_totales': litros_totales,
        'ingresos_leche': ingresos_leche,
        'ingresos_animales': ingresos_animales,
        'ingresos_totales': ingresos_totales,
        'egresos_totales': egresos_totales,
        'balance_neto': balance_neto,
        'gastos': gastos,
        'chart_gastos': chart_gastos,
    }
    return render(request, 'finanzas.html', context)

def auditoria(request):
    user_id = request.session.get('user')
    if not user_id:
        messages.error(request, 'Debe iniciar sesión primero')
        return redirect('login')
        
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    if not finca_activa_id:
        messages.error(request, 'Debe crear una finca primero')
        return redirect('control')
        
    # Cargar logs de actividad asociados a la finca actual o generales (sin finca)
    logs = LogActividad.objects.filter(finca_id=finca_activa_id).order_by('-fecha_hora')
    
    # Filtros
    modulo_filtro = request.GET.get('modulo')
    accion_filtro = request.GET.get('accion')
    q = request.GET.get('q')
    
    if modulo_filtro:
        logs = logs.filter(modulo=modulo_filtro)
    if accion_filtro:
        logs = logs.filter(accion=accion_filtro)
    if q:
        logs = logs.filter(descripcion__icontains=q)
        
    context = {
        'fincas_usuario': fincas_usuario,
        'logs': logs[:200],  # Mostrar los últimos 200 logs
        'modulo_filtro': modulo_filtro,
        'accion_filtro': accion_filtro,
        'q': q,
    }
    return render(request, 'auditoria.html', context)

def perfil(request):
    user_id = request.session.get('user')
    if not user_id:
        messages.error(request, 'Debe iniciar sesión primero')
        return redirect('login')
        
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    
    # Obtener suscripción de app2
    suscripcion = None
    try:
        from app2.models import Suscripcion
        suscripcion, _ = Suscripcion.objects.get_or_create(usuario=user)
    except Exception:
        pass
        
    config, _ = ConfiguracionUsuario.objects.get_or_create(user=user)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'actualizar_datos':
            nombre = request.POST.get('nombre')
            email = request.POST.get('email')
            
            if User.objects.exclude(id=user.id).filter(nombre=nombre).exists():
                messages.error(request, 'El nombre de usuario ya está en uso')
            elif email and User.objects.exclude(id=user.id).filter(email=email).exists():
                messages.error(request, 'El correo electrónico ya está en uso')
            else:
                user.nombre = nombre
                user.email = email
                user.save()
                registrar_log(user, finca_activa_id, 'MODIFICACION', 'SEGURIDAD', f"Actualizó sus datos básicos de cuenta (Nombre: {nombre})")
                messages.success(request, 'Datos de cuenta actualizados correctamente')
                
        elif action == 'cambiar_password':
            current_pw = request.POST.get('current_password')
            new_pw = request.POST.get('new_password')
            confirm_pw = request.POST.get('confirm_password')
            
            if not check_password(current_pw, user.password) and not (not user.password.startswith('pbkdf2_') and user.password == current_pw):
                messages.error(request, 'La contraseña actual es incorrecta')
            elif new_pw != confirm_pw:
                messages.error(request, 'La nueva contraseña y su confirmación no coinciden')
            else:
                user.password = make_password(new_pw)
                user.save()
                registrar_log(user, finca_activa_id, 'MODIFICACION', 'SEGURIDAD', "Actualizó su contraseña de acceso")
                messages.success(request, 'Contraseña cambiada correctamente')
                
        elif action == 'guardar_config':
            config.usar_mamanto = 'usar_mamanto' in request.POST
            config.usar_destete = 'usar_destete' in request.POST
            config.meses_mamanto = int(request.POST.get('meses_mamanto', 3))
            config.meses_destete = int(request.POST.get('meses_destete', 7))
            config.save()
            registrar_log(user, finca_activa_id, 'CONFIGURACION', 'SEGURIDAD', "Modificó los parámetros de crianza y destete")
            messages.success(request, 'Configuración de crianza guardada')
            
        return redirect('perfil')
        
    context = {
        'fincas_usuario': fincas_usuario,
        'user_profile': user,
        'suscripcion': suscripcion,
        'config': config,
    }
    return render(request, 'perfil.html', context)

def engorde(request):
    user_id = request.session.get('user')
    if not user_id:
        messages.error(request, 'Debe iniciar sesión primero')
        return redirect('login')
        
    user = User.objects.get(id=user_id)
    finca_activa_id, fincas_usuario = get_finca_context(request, user)
    if not finca_activa_id:
        messages.error(request, 'Debe crear una finca primero')
        return redirect('control')
        
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'crear_corral':
            nombre = request.POST.get('nombre')
            capacidad = request.POST.get('capacidad', 20)
            desc = request.POST.get('descripcion', '')
            try:
                Corral.objects.create(finca_id=finca_activa_id, nombre=nombre, capacidad=capacidad, descripcion=desc)
                registrar_log(user, finca_activa_id, 'CREACION', 'REBAÑOS', f"Creó corral: '{nombre}' con capacidad para {capacidad} animales")
                messages.success(request, 'Corral creado correctamente')
            except Exception as e:
                messages.error(request, f'Error al crear corral: {str(e)}')
                
        elif action == 'registrar_alimentacion':
            corral_id = request.POST.get('corral_id')
            alimento = request.POST.get('tipo_alimento')
            cantidad = request.POST.get('cantidad_kg')
            costo = request.POST.get('costo_total', 0.00)
            fecha = request.POST.get('fecha')
            try:
                corral = Corral.objects.get(id=corral_id, finca_id=finca_activa_id)
                RegistroAlimentacion.objects.create(
                    finca_id=finca_activa_id,
                    corral=corral,
                    fecha=fecha,
                    tipo_alimento=alimento,
                    cantidad_kg=cantidad,
                    costo_total=costo
                )
                registrar_log(user, finca_activa_id, 'CREACION', 'FINANZAS', f"Registró alimentación en corral '{corral.nombre}': {cantidad}kg de {alimento} por ${costo}")
                messages.success(request, 'Alimentación registrada con éxito')
            except Exception as e:
                messages.error(request, f'Error al registrar alimentación: {str(e)}')
                
        elif action == 'registrar_pesaje':
            animal_id = request.POST.get('animal_id')
            peso = request.POST.get('peso_kg')
            fecha = request.POST.get('fecha')
            try:
                animal = Animal.objects.get(id=animal_id, finca_id=finca_activa_id)
                PesajeAnimal.objects.create(animal=animal, fecha=fecha, peso_kg=peso)
                registrar_log(user, finca_activa_id, 'CREACION', 'ANIMALES', f"Registró pesaje para el animal {animal.codigo}: {peso}kg")
                messages.success(request, f'Pesaje de {peso}kg registrado para el animal {animal.codigo}')
            except Exception as e:
                messages.error(request, f'Error al registrar pesaje: {str(e)}')
                
        elif action == 'crear_tarea':
            desc = request.POST.get('descripcion')
            cat = request.POST.get('categoria', 'OTRO')
            fecha = request.POST.get('fecha')
            try:
                TareaDiaria.objects.create(finca_id=finca_activa_id, fecha=fecha, descripcion=desc, categoria=cat)
                messages.success(request, 'Tarea diaria programada')
            except Exception as e:
                messages.error(request, f'Error al crear tarea: {str(e)}')
                
        elif action == 'toggle_tarea':
            tarea_id = request.POST.get('tarea_id')
            try:
                tarea = TareaDiaria.objects.get(id=tarea_id, finca_id=finca_activa_id)
                tarea.completada = not tarea.completada
                tarea.save()
                messages.success(request, 'Estado de la tarea actualizado')
            except Exception as e:
                messages.error(request, f'Error al actualizar tarea: {str(e)}')
                
        elif action == 'transferir_animales':
            destino_id = request.POST.get('corral_destino_id')
            animal_ids = request.POST.getlist('animal_ids')
            motivo = request.POST.get('motivo', '')
            try:
                destino = Corral.objects.get(id=destino_id, finca_id=finca_activa_id) if destino_id else None
                for a_id in animal_ids:
                    animal = Animal.objects.get(id=a_id, finca_id=finca_activa_id)
                    origen = animal.corral
                    animal.corral = destino
                    animal.save()
                    HistorialTransferencia.objects.create(
                        animal=animal,
                        corral_origen=origen,
                        corral_destino=destino,
                        motivo=motivo
                    )
                dest_nombre = destino.nombre if destino else 'Ninguno'
                registrar_log(user, finca_activa_id, 'MODIFICACION', 'REBAÑOS', f"Transfirió {len(animal_ids)} animal(es) al corral '{dest_nombre}'")
                messages.success(request, 'Animales transferidos correctamente')
            except Exception as e:
                messages.error(request, f'Error en transferencia: {str(e)}')
                
        return redirect('engorde')
        
    # GET
    corrales_list = Corral.objects.filter(finca_id=finca_activa_id)
    corrales_data = []
    
    for c in corrales_list:
        animales = c.animales.filter(estado_vida='VIVO')
        
        # Calcular GDP (Ganancia Diaria Promedio) del corral
        total_gdp = 0.0
        con_gdp = 0
        for a in animales:
            pesajes = a.pesajes.all().order_by('fecha')
            if pesajes.count() >= 2:
                p_primero = pesajes.first()
                p_ultimo = pesajes.last()
                dias = (p_ultimo.fecha - p_primero.fecha).days
                if dias > 0:
                    gdp = float(p_ultimo.peso_kg - p_primero.peso_kg) / dias
                    total_gdp += gdp
                    con_gdp += 1
                    
        gdp_promedio = total_gdp / con_gdp if con_gdp > 0 else 0.0
        costo_alimento = RegistroAlimentacion.objects.filter(corral=c).aggregate(total=Sum('costo_total'))['total'] or 0
        
        corrales_data.append({
            'corral': c,
            'animales': animales,
            'gdp_promedio': gdp_promedio,
            'costo_alimento': costo_alimento,
            'ocupacion_porcentaje': (animales.count() / c.capacidad * 100) if c.capacidad > 0 else 0,
        })
        
    animales_sin_corral = Animal.objects.filter(finca_id=finca_activa_id, estado_vida='VIVO', corral__isnull=True)
    animales_todos = Animal.objects.filter(finca_id=finca_activa_id, estado_vida='VIVO')
    
    hoy = datetime.date.today()
    tareas_pendientes = TareaDiaria.objects.filter(finca_id=finca_activa_id, fecha=hoy).order_by('completada')
    
    context = {
        'fincas_usuario': fincas_usuario,
        'corrales': corrales_data,
        'animales_sin_corral': animales_sin_corral,
        'animales_todos': animales_todos,
        'tareas': tareas_pendientes,
    }
    return render(request, 'engorde.html', context)
